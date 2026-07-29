"""Export gamma_fusion results to full-metric Excel + Markdown (SOTA bold)."""
from __future__ import annotations

import csv
import json
import os
from os.path import join as j

import numpy as np

import re

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

TASK_RE = re.compile(r'_task(\d+)_')
TASK_LABELS = {
    '13': 'CN vs AD (阳性=AD)',
    '23': 'MCI vs AD (阳性=AD)',
    '12': 'CN vs MCI / NC vs MCI (阳性=MCI)',
}


def _parse_class_task(tag: str) -> str:
    m = TASK_RE.search(tag or '')
    return m.group(1) if m else '13'


ORDER = [
    'B0', 'B_gamma', 'B_film', 'B_daft', 'B_hyper', 'B_concat',
    'L0', 'L1', 'L2', 'L3',
    'A_global', 'A_noshare', 'A_shuffle',
]

ID_KEYS = [
    ('B0', 'gamma_fusion_B0'),
    ('B_gamma', 'gamma_fusion_B_gamma'),
    ('B_film', 'gamma_fusion_B_film'),
    ('B_daft', 'gamma_fusion_B_daft'),
    ('B_hyper', 'gamma_fusion_B_hyper'),
    ('B_concat', 'gamma_fusion_B_concat'),
    ('L0', 'gamma_fusion_L0'),
    ('L1', 'gamma_fusion_L1'),
    ('L2', 'gamma_fusion_L2'),
    ('L3', 'gamma_fusion_L3'),
    ('A_global', 'gamma_fusion_A_global'),
    ('A_noshare', 'gamma_fusion_A_noshare'),
    ('A_shuffle', 'gamma_fusion_A_shuffle'),
]

META = {
    'B0': ('Wave0 下界', 'ResNet CE-only（无表格）'),
    'B_gamma': ('Wave0 主方法', 'ResNet + patch-γ（默认正则）'),
    'B_film': ('Wave0 对照', 'ResNet + FiLM（last stage）'),
    'B_daft': ('Wave0 对照', 'ResNet + DAFT'),
    'B_hyper': ('Wave0 对照', 'HyperFusion'),
    'B_concat': ('Wave0 对照', 'Concat 晚融合'),
    'L0': ('Wave1 损失', 'γ + CE only（无 sparsity/smooth/dropout）'),
    'L1': ('Wave1 损失', 'γ + sparsity only'),
    'L2': ('Wave1 损失', 'γ + smoothness only'),
    'L3': ('Wave1 损失', '同 B_gamma（完整正则，别名）'),
    'A_global': ('Wave2 架构', '全局标量 γ'),
    'A_noshare': ('Wave2 架构', 'patch γ，无 group sharing'),
    'A_shuffle': ('Wave2 架构', 'batch 内打乱表格'),
}

# Head-to-head pool for Wave0 SOTA claim
WAVE0_IDS = {'B0', 'B_gamma', 'B_film', 'B_daft', 'B_hyper', 'B_concat'}

METRIC_KEYS = [
    ('val_acc', 'Val Acc', True),
    ('test_acc', 'Acc', True),
    ('test_sen', 'Sen', True),
    ('test_f1', 'F1', True),
    ('test_auc', 'AUC', True),
    ('test_spe', 'Spe', True),
    ('test_loss', 'Loss', False),  # lower is better
]


def _mean_std(vals):
    a = np.asarray([float(v) for v in vals], dtype=np.float64)
    return float(a.mean()), float(a.std())


def _pm(mean, std, digits=3):
    if mean is None:
        return '—'
    if std is None:
        return f'{mean:.{digits}f}'
    return f'{mean:.{digits}f} ± {std:.{digits}f}'


def _g(v, digits=3):
    if v is None:
        return '—'
    return f'{float(v):.{digits}f}'


def _parse_pm_token(s: str):
    s = (s or '').strip().replace('**', '')
    if not s or s in ('—', '-', 'N/A'):
        return None, None
    if '±' in s:
        a, b = s.split('±', 1)
        return float(a.strip()), float(b.strip())
    return float(s), None


def parse_5090_md(path: str) -> dict[str, dict]:
    """Parse weights/5090.md (task13 on RTX 5090 server) into metric dicts by exp ID."""
    if not os.path.isfile(path):
        return {}
    text = open(path, encoding='utf-8').read()
    metrics: dict[str, dict] = {}

    main_cols = {
        3: 'test_acc', 4: 'test_sen', 5: 'test_f1',
        6: 'test_auc', 7: 'test_spe',
    }
    in_main = False
    for line in text.splitlines():
        if line.startswith('## 2.'):
            in_main = True
            continue
        if in_main and line.startswith('### 附加'):
            break
        if not in_main or not line.startswith('|'):
            continue
        if 'ID' in line and 'Acc' in line:
            continue
        if set(line.replace('|', '').replace('-', '').strip()) <= set():
            continue
        parts = [p.strip().replace('**', '') for p in line.strip().strip('|').split('|')]
        if len(parts) < 8:
            continue
        eid = parts[0]
        if eid not in ORDER or eid == 'L3':
            continue
        row = metrics.setdefault(eid, {'n_folds': 5})
        for idx, key in main_cols.items():
            m, s = _parse_pm_token(parts[idx])
            row[f'{key}_mean'] = m
            row[f'{key}_std'] = s

    in_extra = False
    for line in text.splitlines():
        if '### 附加' in line:
            in_extra = True
            continue
        if in_extra and line.startswith('## ') and '附加' not in line:
            break
        if not in_extra or not line.startswith('|'):
            continue
        if 'ID' in line or set(line.replace('|', '').replace('-', '').strip()) <= set():
            continue
        parts = [p.strip().replace('**', '') for p in line.strip().strip('|').split('|')]
        if len(parts) < 3:
            continue
        eid = parts[0]
        if eid not in metrics:
            continue
        m, s = _parse_pm_token(parts[1])
        metrics[eid]['val_acc_mean'] = m
        metrics[eid]['val_acc_std'] = s
        m, s = _parse_pm_token(parts[2])
        metrics[eid]['test_loss_mean'] = m
        metrics[eid]['test_loss_std'] = s

    if 'B_gamma' in metrics and 'L3' not in metrics:
        metrics['L3'] = dict(metrics['B_gamma'])
        metrics['L3']['run_tag'] = '5090_server_task13_L3_alias'
    for eid, row in metrics.items():
        row.setdefault('run_tag', f'5090_server_task13_{eid}')
        row['gamma_mean'] = None
        row['corr_gamma_age'] = None
    return metrics


def _apply_sota_flags(rows):
    def best_id(candidates):
        best_eid, best_auc = None, None
        for r in rows:
            if r['ID'] not in candidates:
                continue
            auc = r.get('test_auc_mean')
            if auc is None:
                continue
            if best_auc is None or auc > best_auc + 1e-12:
                best_eid, best_auc = r['ID'], auc
        return best_eid

    w0 = best_id(WAVE0_IDS)
    overall = best_id({r['ID'] for r in rows if r['ID'] != 'L3'})
    for r in rows:
        r['is_sota_wave0'] = r['ID'] == w0
        r['is_sota_overall'] = r['ID'] == overall
        r['is_sota'] = r['is_sota_wave0'] or r['is_sota_overall']
        marks = []
        if r['is_sota_wave0']:
            marks.append('Wave0-SOTA')
        if r['is_sota_overall']:
            marks.append('Overall-best AUC')
        r['SOTA'] = ' / '.join(marks) if marks else ''


def _resolve_exp_id(tag: str):
    for eid, key in ID_KEYS:
        if key in tag:
            return eid
    return None


def _read_fold_tsv(run_dir):
    path = j(run_dir, 'summary', 'test_per_fold.tsv')
    if not os.path.isfile(path):
        return []
    with open(path, encoding='utf-8') as f:
        return list(csv.DictReader(f, delimiter='\t'))


def _read_gamma_stats(run_dir):
    means, corrs = [], []
    for name in sorted(os.listdir(run_dir)):
        if not name.startswith('fold_'):
            continue
        gp = j(run_dir, name, 'test', 'gamma_stats.json')
        if not os.path.isfile(gp):
            continue
        with open(gp, encoding='utf-8') as f:
            gs = json.load(f)
        if gs.get('gamma_mean') is not None:
            means.append(float(gs['gamma_mean']))
        if gs.get('corr_gamma_age') is not None:
            corrs.append(float(gs['corr_gamma_age']))
    return (
        float(np.mean(means)) if means else None,
        float(np.mean(corrs)) if corrs else None,
    )


def discover_runs(roots):
    found = {}
    for root in roots:
        if not os.path.isdir(root):
            continue
        for name in sorted(os.listdir(root)):
            path = j(root, name)
            if not os.path.isdir(path) or 'gamma_fusion' not in name:
                continue
            eid = _resolve_exp_id(name)
            if not eid:
                continue
            class_task = _parse_class_task(name)
            folds = _read_fold_tsv(path)
            if len(folds) < 1:
                continue
            row = {
                'ID': eid,
                'class_task': class_task,
                'run_dir': path,
                'run_tag': name,
                'n_folds': len(folds),
            }
            for key, _label, _higher in METRIC_KEYS:
                try:
                    vals = [float(f[key]) for f in folds]
                except (KeyError, ValueError):
                    row[f'{key}_mean'] = None
                    row[f'{key}_std'] = None
                    continue
                m, s = _mean_std(vals)
                row[f'{key}_mean'] = m
                row[f'{key}_std'] = s
            gm, cr = _read_gamma_stats(path)
            row['gamma_mean'] = gm
            row['corr_gamma_age'] = cr
            found[(class_task, eid)] = row

    # L3 alias per task
    for task in sorted({k[0] for k in found}):
        if (task, 'B_gamma') in found and (task, 'L3') not in found:
            alias = dict(found[(task, 'B_gamma')])
            alias['ID'] = 'L3'
            alias['run_tag'] = (alias.get('run_tag') or '') + '_alias_L3'
            found[(task, 'L3')] = alias
    return found


def _row_from_raw(eid: str, r: dict, source: str = 'local') -> dict:
    wave, desc = META.get(eid, ('', ''))
    out = {
        '来源': source,
        'ID': eid,
        '分组': wave,
        '方法说明': desc,
        '折数': r.get('n_folds') or '',
        'run_tag': r.get('run_tag') or '',
        'is_sota_wave0': False,
        'is_sota_overall': False,
    }
    for key, label, _ in METRIC_KEYS:
        m, s = r.get(f'{key}_mean'), r.get(f'{key}_std')
        out[label] = _pm(m, s)
        out[f'{key}_mean'] = m
        out[f'{key}_std'] = s
    out['γ mean'] = _g(r.get('gamma_mean'))
    out['corr(γ, age)'] = _g(r.get('corr_gamma_age'))
    out['gamma_mean'] = r.get('gamma_mean')
    out['corr_gamma_age'] = r.get('corr_gamma_age')
    return out


def build_rows(found, class_task: str = '13', source: str = 'local'):
    rows = []
    for eid in ORDER:
        r = found.get((class_task, eid), {})
        if not r:
            continue
        rows.append(_row_from_raw(eid, r, source=source))
    _apply_sota_flags(rows)
    return rows


def build_rows_from_external(metrics_by_id: dict[str, dict], source: str = '5090'):
    rows = []
    for eid in ORDER:
        r = metrics_by_id.get(eid)
        if not r:
            continue
        rows.append(_row_from_raw(eid, r, source=source))
    _apply_sota_flags(rows)
    return rows


DISPLAY_COLS = [
    '来源', 'ID', '分组', '方法说明', 'SOTA', '折数',
    'Val Acc', 'Acc', 'Sen', 'F1', 'AUC', 'Spe', 'Loss',
    'γ mean', 'corr(γ, age)',
]


def write_markdown(rows, path):
    lines = [
        '# ResNet+γ vs 表格融合：std5cv 全指标主表',
        '',
        '协议：ADNI CN/AD，受试者级 5-fold（~60/20/20）；指标为 **held-out test** mean ± std（Val Acc 为选模用验证集）。',
        '',
        '> **Sen / Spe**：以阳性类 **AD** 为基准的标准二分类定义 '
        'Sen=TP/(TP+FN)，Spe=TN/(TN+FP)（非 multiclass-macro；二者不再恒等）。',
        '',
        '> **加粗** = Wave0 头对头 SOTA（相对 FiLM/DAFT/Hyper/Concat）和/或 全表最高 test AUC。',
        '',
        '| ' + ' | '.join(DISPLAY_COLS) + ' |',
        '| ' + ' | '.join(['---'] * len(DISPLAY_COLS)) + ' |',
    ]
    for r in rows:
        cells = []
        for c in DISPLAY_COLS:
            v = str(r.get(c, ''))
            if r.get('is_sota'):
                v = f'**{v}**'
            cells.append(v)
        lines.append('| ' + ' | '.join(cells) + ' |')
    lines.extend(['', ''])
    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print('Wrote', path)


def write_markdown_section(rows, lines_out: list):
    lines_out.append('| ' + ' | '.join(DISPLAY_COLS) + ' |')
    lines_out.append('| ' + ' | '.join(['---'] * len(DISPLAY_COLS)) + ' |')
    for r in rows:
        cells = []
        for c in DISPLAY_COLS:
            v = str(r.get(c, ''))
            if r.get('is_sota'):
                v = f'**{v}**'
            cells.append(v)
        lines_out.append('| ' + ' | '.join(cells) + ' |')


def write_excel_multitask(all_rows_by_task: dict, path: str):
    try:
        from openpyxl import Workbook
    except ImportError:
        import subprocess
        import sys
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
        from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for task in sorted(all_rows_by_task.keys()):
        rows = all_rows_by_task[task]
        ws = wb.create_sheet(title=f'task{task}'[:31])
        _fill_excel_sheet(ws, rows)
    ws_note = wb.create_sheet(title='说明')
    notes = [
        '多任务 gamma_fusion std5cv；table_feature=1 → AGE_YEARS only',
        '来源=local：本机 weights/classifier/age_cv_summary/ADNI',
        '来源=5090：RTX 5090 服务器 task13 结果（见 weights/5090.md）',
        'SOTA 在各自来源内独立计算（Wave0 头对头 / Overall-best AUC）',
    ]
    for i, line in enumerate(notes, 1):
        ws_note.cell(i, 1, line)
    ws_note.column_dimensions['A'].width = 100
    wb.save(path)
    print('Wrote', path)


def _fill_excel_sheet(ws, rows):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(color='FFFFFF', bold=True)
    sota_fill = PatternFill('solid', fgColor='C6EFCE')
    thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    wave_fills = {
        'Wave0 下界': PatternFill('solid', fgColor='F2F2F2'),
        'Wave0 主方法': PatternFill('solid', fgColor='E2EFDA'),
        'Wave0 对照': PatternFill('solid', fgColor='DDEBF7'),
        'Wave1 损失': PatternFill('solid', fgColor='FFF2CC'),
        'Wave2 架构': PatternFill('solid', fgColor='FCE4D6'),
    }
    for c, h in enumerate(DISPLAY_COLS, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    for i, r in enumerate(rows, 2):
        base_fill = sota_fill if r.get('is_sota') else wave_fills.get(r['分组'])
        for c, h in enumerate(DISPLAY_COLS, 1):
            cell = ws.cell(i, c, r.get(h, ''))
            cell.border = thin
            if base_fill:
                cell.fill = base_fill
            if r.get('is_sota'):
                cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'


def write_excel(rows, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess
        import sys
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = '主表'

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(color='FFFFFF', bold=True)
    sota_fill = PatternFill('solid', fgColor='C6EFCE')
    thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    wave_fills = {
        'Wave0 下界': PatternFill('solid', fgColor='F2F2F2'),
        'Wave0 主方法': PatternFill('solid', fgColor='E2EFDA'),
        'Wave0 对照': PatternFill('solid', fgColor='DDEBF7'),
        'Wave1 损失': PatternFill('solid', fgColor='FFF2CC'),
        'Wave2 架构': PatternFill('solid', fgColor='FCE4D6'),
    }

    for c, h in enumerate(DISPLAY_COLS, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin

    for i, r in enumerate(rows, 2):
        base_fill = sota_fill if r.get('is_sota') else wave_fills.get(r['分组'])
        bold = Font(bold=True)
        for c, h in enumerate(DISPLAY_COLS, 1):
            cell = ws.cell(i, c, r.get(h, ''))
            cell.border = thin
            cell.alignment = Alignment(
                horizontal='left' if h in ('方法说明', 'SOTA') else 'center',
                vertical='center',
                wrap_text=True,
            )
            if base_fill:
                cell.fill = base_fill
            if r.get('is_sota'):
                cell.font = bold

    widths = {
        'ID': 12, '分组': 12, '方法说明': 34, 'SOTA': 18, '折数': 6,
        'Val Acc': 14, 'Acc': 14, 'Sen': 14, 'F1': 14, 'AUC': 14,
        'Spe': 14, 'Loss': 14, 'γ mean': 10, 'corr(γ, age)': 12,
    }
    for i, h in enumerate(DISPLAY_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 12)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(DISPLAY_COLS))}{len(rows) + 1}'

    # Numeric sheet
    ws2 = wb.create_sheet('数值')
    num_headers = ['ID', '分组', 'SOTA']
    for key, label, _ in METRIC_KEYS:
        num_headers += [f'{label}_mean', f'{label}_std']
    num_headers += ['gamma_mean', 'corr_gamma_age', 'run_tag']
    for c, h in enumerate(num_headers, 1):
        cell = ws2.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    for i, r in enumerate(rows, 2):
        vals = {
            'ID': r['ID'], '分组': r['分组'], 'SOTA': r['SOTA'],
            'gamma_mean': r.get('gamma_mean'),
            'corr_gamma_age': r.get('corr_gamma_age'),
            'run_tag': r.get('run_tag'),
        }
        for key, label, _ in METRIC_KEYS:
            vals[f'{label}_mean'] = r.get(f'{key}_mean')
            vals[f'{label}_std'] = r.get(f'{key}_std')
        for c, h in enumerate(num_headers, 1):
            cell = ws2.cell(i, c, vals.get(h))
            cell.border = thin
            if isinstance(vals.get(h), float):
                cell.number_format = '0.000'
            if r.get('is_sota'):
                cell.font = Font(bold=True)
                cell.fill = sota_fill
    for i in range(1, len(num_headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 12
    ws2.freeze_panes = 'A2'

    ws3 = wb.create_sheet('说明')
    notes = [
        'ResNet+γ vs 表格融合（std5cv，无 Phase1）',
        '',
        '全指标：Val Acc / Acc / Sen / F1 / AUC / Spe / Loss（mean±std over 5 folds）',
        'γ mean、corr(γ, age) 来自每折 test/gamma_stats.json（仅 backdoor 有）',
        '',
        '加粗 + 绿色底：SOTA',
        '  - Wave0-SOTA：头对头集合 {B0,B_gamma,B_film,B_daft,B_hyper,B_concat} 中最高 test AUC',
        '  - Overall-best AUC：全表（不含 L3 别名）最高 test AUC',
        '',
        '协议：ADNI task13；table_feature=1 → AGE_YEARS only',
        '复盘：utils/review/gamma_vs_fusion_5fold_review.md',
    ]
    for i, line in enumerate(notes, 1):
        ws3.cell(i, 1, line)
    ws3.column_dimensions['A'].width = 100

    wb.save(path)
    print('Wrote', path)


def main():
    roots = [j(PROJECT_ROOT, 'weights', 'classifier', 'age_cv_summary', 'ADNI')]
    found = discover_runs(roots)
    ext_5090_path = j(PROJECT_ROOT, 'weights', '5090.md')
    ext_5090 = parse_5090_md(ext_5090_path)
    tasks = sorted({k[0] for k in found}) or ['13']
    if ext_5090 and '13' not in tasks:
        tasks.append('13')
        tasks = sorted(set(tasks))
    out_dir = j(PROJECT_ROOT, 'weights', 'classifier')

    md_parts = [
        '# ResNet+γ vs 表格融合：多任务 std5cv 全指标主表',
        '',
        '协议：ADNI；受试者级 5-fold（~60/20/20）；`table_feature=1` → **AGE_YEARS only**。',
        'Sen/Spe：task13/23 阳性=AD；task12 阳性=MCI。',
        '',
    ]
    all_rows_by_task = {}
    for task in tasks:
        rows = build_rows(found, class_task=task, source='local')
        if task == '13' and ext_5090:
            rows = rows + build_rows_from_external(ext_5090, source='5090')
        if not rows:
            continue
        all_rows_by_task[task] = rows
        label = TASK_LABELS.get(task, f'task{task}')
        md_parts.append(f'## Task {task}: {label}')
        md_parts.append('')
        write_markdown_section(rows, md_parts)
        md_parts.append('')

    multitask_md = j(out_dir, 'gamma_fusion_multitask_summary.md')
    with open(multitask_md, 'w', encoding='utf-8') as f:
        f.write('\n'.join(md_parts))
    print('Wrote', multitask_md)

    # Per-task + legacy single-task md for task13
    if '13' in all_rows_by_task:
        write_markdown(all_rows_by_task['13'], j(out_dir, 'gamma_fusion_5fold_summary.md'))

    xlsx = j(out_dir, 'gamma_fusion_multitask_summary.xlsx')
    try:
        write_excel_multitask(all_rows_by_task, xlsx)
    except PermissionError:
        alt = j(out_dir, 'gamma_fusion_multitask_summary_new.xlsx')
        write_excel_multitask(all_rows_by_task, alt)
        print(f'NOTE: {xlsx} locked; wrote {alt}')

    # Machine-readable full TSV (all tasks)
    tsv_path = j(out_dir, 'gamma_fusion_multitask_summary_full.tsv')
    fields = [
        'class_task', '来源', 'ID', '分组', '方法说明', 'SOTA', '折数',
        'val_acc_mean', 'val_acc_std',
        'test_acc_mean', 'test_acc_std',
        'test_sen_mean', 'test_sen_std',
        'test_f1_mean', 'test_f1_std',
        'test_auc_mean', 'test_auc_std',
        'test_spe_mean', 'test_spe_std',
        'test_loss_mean', 'test_loss_std',
        'gamma_mean', 'corr_gamma_age', 'run_tag',
    ]
    with open(tsv_path, 'w', encoding='utf-8', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter='\t', extrasaction='ignore')
        w.writeheader()
        for task, rows in sorted(all_rows_by_task.items()):
            for r in rows:
                r2 = dict(r)
                r2['class_task'] = task
                w.writerow(r2)
    print('Wrote', tsv_path)


if __name__ == '__main__':
    main()
