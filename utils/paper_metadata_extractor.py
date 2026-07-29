# -*- coding: utf-8 -*-
"""
论文元数据提取工具（PDF → Excel）

方法总结
========
本模块从 MR 加速相关 PDF 的前几页自动提取结构化信息，无需通读全文。
核心思路是「按出版商类型分流解析 + 首页版式启发式规则」。

1. 出版商分流
   - IEEE TMI / JBHI：标题在 VOL 头之后、Abstract 之前；作者行通常含逗号或
     "Member, IEEE" 等后缀；首作者可能单独占一行（无逗号），下一行以 "," 开头。
   - Elsevier Media：标题在 "locate/media" 或期刊 boilerplate 之后；作者以
     "Name a," / "a, Name" 等形式标注机构；需区分作者标记行与真正的机构地址行。
   - MICCAI：标题若干行后接一行 "Name1,2, Name2, and Name3" 格式。
   - ISBI：全大写标题后，每行一个 "Name12" 或 "Name Name" 作者。

2. 标题提取
   - 优先使用描述性文件名（TMI/JBHI/ISBI 子目录下的 PDF）。
   - Elsevier 加密文件名（1-s2.0-*.pdf）从首页 parse；同行附带作者时截断。
   - MICCAI 编号文件名（1256_paper.pdf）合并多行标题直至作者行。

3. 作者提取（重点）
   - 定位 Abstract：支持 "Abstract", "ABSTRACT", "A B S T R A C T" 等变体。
   - IEEE：收集 Abstract 前含逗号 / IEEE 会员标记 / 孤立姓名行；去除标题前缀；
     清除 "Member, IEEE" 等待遇后缀。
   - Elsevier：状态机 skip_title → authors；作者行含 "Name a," 或纯姓名行；
     "a, Guanxiong Luo" 视为作者续行而非机构；去除 affiliation 字母 a/b/c。
   - 姓名规范化：去数字上标、*†、尾部 " a"、"(B)" 等；过滤含 reconstruction 等
     论文关键词的误识别片段。

4. 其他字段
   - 年份：IEEE VOL 行、Elsevier 卷期行、文件名中的年份。
   - URL：DOI / IEEE DOI / arXiv 正则匹配。
   - 代码：前 5 页搜索 github.com 链接。
   - 数据集（分层提取，见 extract_datasets）：
     * 已知目录：fastMRI、CMRxRecon、ADNI 等 30+ 公开 MRI 数据集；
     * 章节聚焦：定位 Dataset/Experiments/Methods 等节提取上下文；
     * 模式发现：匹配 "X dataset/database/benchmark/challenge" 等句式发现新数据集；
     * URL 关联：Zenodo、Synapse、Grand-Challenge 及 "available at" 链接与数据集名绑定。

依赖：pymupdf (fitz), openpyxl

用法示例
--------
>>> from utils.paper_metadata_extractor import extract_paper_info, fill_summary_xlsx
>>> info = extract_paper_info(Path("paper.pdf"))
>>> fill_summary_xlsx(Path("summary.xlsx"), Path("papers_root"))
"""
from __future__ import annotations

import json
import re
import shutil
import time
import urllib.error
import urllib.request
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import fitz  # pymupdf
import openpyxl

_ILLEGAL_XL_RE = re.compile(r"[\000-\010\013\014\016-\037\ufffd\ufeff]")

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------

FOLDER_VENUE = {
    "ISBI 2026": "IEEE ISBI 2026",
    "JBHI": "IEEE Journal of Biomedical and Health Informatics (JBHI)",
    "Media": "Medical Image Analysis",
    "MICCAI 2025": "MICCAI 2025",
    "TMI": "IEEE Transactions on Medical Imaging (TMI)",
    "脊柱MRI超分": "脊柱MRI超分",
    "模型驱动": "模型驱动",
    "WSI多目标分割": "WSI多目标分割",
    "no relative": "no relative",
}

DEFAULT_FOLDER_ORDER = [
    "TMI", "Media", "JBHI", "MICCAI 2025", "ISBI 2026",
    "脊柱MRI超分", "模型驱动", "WSI多目标分割", "no relative",
]

# PDF 首页常见期刊/会议 → 规范名称
VENUE_PATTERNS: list[tuple[str, str]] = [
    (r"IEEE TRANSACTIONS ON MEDICAL IMAGING", "IEEE Transactions on Medical Imaging (TMI)"),
    (r"IEEE JOURNAL OF BIOMEDICAL AND HEALTH INFORMATICS", "IEEE Journal of Biomedical and Health Informatics (JBHI)"),
    (r"IEEE TRANSACTIONS ON CIRCUITS AND SYSTEMS FOR VIDEO TECHNOLOGY", "IEEE Transactions on Circuits and Systems for Video Technology (TCSVT)"),
    (r"IEEE TRANSACTIONS ON IMAGE PROCESSING", "IEEE Transactions on Image Processing (TIP)"),
    (r"IEEE TRANSACTIONS ON NEURAL NETWORKS AND LEARNING SYSTEMS", "IEEE Transactions on Neural Networks and Learning Systems (TNNLS)"),
    (r"International Symposium on Biomedical Imaging|\bISBI\s+\d{4}\b|\bISBI\b", "IEEE ISBI"),
    (r"\bMICCAI\b", "MICCAI"),
    (r"Medical Image Analysis", "Medical Image Analysis"),
    (r"NeuroImage|Neuroimage", "NeuroImage"),
    (r"Magnetic Resonance in Medicine|\bMRM\b", "Magnetic Resonance in Medicine (MRM)"),
    (r"Scientific Reports", "Scientific Reports"),
    (r"BMC Medicine", "BMC Medicine"),
    (r"Progress in Biomedical Engineering", "Progress in Biomedical Engineering"),
    (r"Journal of Healthcare Informatics Research", "Journal of Healthcare Informatics Research"),
    (r"Advances in Neural Information Processing Systems|\bNeurIPS\b", "NeurIPS"),
    (r"\barXiv:\d{4}\.\d+", "arXiv"),
    (r"Frontiers in", "Frontiers"),
    (r"Radiology", "Radiology"),
    (r"European Radiology", "European Radiology"),
    (r"Journal of Magnetic Resonance Imaging|\bJMRI\b", "Journal of Magnetic Resonance Imaging (JMRI)"),
    (r"Nature Medicine|Nature Methods|Nature Communications|\bNature\b", "Nature 系列"),
    (r"Computer Vision and Pattern Recognition|\bCVPR\b", "IEEE CVPR"),
    (r"International Conference on Computer Vision|\bICCV\b", "IEEE ICCV"),
    (r"European Conference on Computer Vision|\bECCV\b", "ECCV"),
    (r"International Conference on Medical Image Computing|\bMICCAI\b", "MICCAI"),
    (r"Histopathology|Modern Pathology|Journal of Pathology", "病理学期刊"),
    (r"IEEE Access", "IEEE Access"),
]

# 扩展已知 MRI / 医学影像数据集目录（pattern → canonical name → url）
KNOWN_DATASETS: list[dict[str, str]] = [
    {"pattern": r"\bfastMRI\b", "name": "fastMRI", "url": "https://fastmri.med.nyu.edu/"},
    {"pattern": r"\bCMRxRecon(?:2024)?\b", "name": "CMRxRecon", "url": "https://cmrxrecon2024.grand-challenge.org/"},
    {"pattern": r"\bCMRxMotion\b", "name": "CMRxMotion", "url": "https://cmrxmotion2022.grand-challenge.org/"},
    {"pattern": r"\bSKM-TEA\b", "name": "SKM-TEA", "url": ""},
    {"pattern": r"\bHuman Connectome Project\b|\bHCP[\-\s]?MMA\b|\bHCP dataset\b", "name": "Human Connectome Project (HCP)", "url": "https://www.humanconnectome.org/"},
    {"pattern": r"\bIXI dataset\b|\bIXI\b(?=[^\n]{0,40}(?:dataset|MRI|T1|T2))", "name": "IXI", "url": "https://brain-development.org/ixi-dataset/"},
    {"pattern": r"\bADNI\b", "name": "ADNI", "url": "https://adni.loni.usc.edu/"},
    {"pattern": r"\bUK Biobank\b", "name": "UK Biobank", "url": "https://www.ukbiobank.ac.uk/"},
    {"pattern": r"\bM4Raw\b", "name": "M4Raw", "url": "https://m4raw.github.io/"},
    {"pattern": r"\bCC359\b|\bCalgary[\-\s]?Campinas\b", "name": "Calgary-Campinas (CC359)", "url": "https://sites.google.com/view/calgary-campinas-dataset/"},
    {"pattern": r"\bOASIS\b", "name": "OASIS", "url": "https://www.oasis-brains.org/"},
    {"pattern": r"\bBraTS\b|\bBRATS\b", "name": "BraTS", "url": "https://www.synapse.org/#!Synapse:syn2580853"},
    {"pattern": r"\bACDC\b(?=[^\n]{0,30}(?:cardiac|MRI|dataset))", "name": "ACDC (cardiac)", "url": "https://www.creatis.insa-lyon.fr/Challenge/acdc/"},
    {"pattern": r"\bMSCMR\b", "name": "MSCMR", "url": ""},
    {"pattern": r"\bMRNet\b", "name": "MRNet", "url": ""},
    {"pattern": r"\bStanford\s*3D\b|\bStanford\s*2D\b", "name": "Stanford 3D/2D FSE", "url": ""},
    {"pattern": r"\bMultiCoil\b", "name": "MultiCoil MRI", "url": ""},
    {"pattern": r"\bProstateX\b", "name": "ProstateX", "url": ""},
    {"pattern": r"\bMSSEG\b", "name": "MSSEG", "url": ""},
    {"pattern": r"\bKirby(?:\s+corpus|\s+dataset)?\b", "name": "Kirby", "url": ""},
    {"pattern": r"\bdHCP\b|\bdeveloping HCP\b", "name": "dHCP", "url": "https://www.developingconnectome.org/"},
    {"pattern": r"\bABCD\b(?=[^\n]{0,30}(?:study|dataset|MRI))", "name": "ABCD", "url": "https://abcdstudy.org/"},
    {"pattern": r"\bCamCAN\b", "name": "CamCAN", "url": "https://camcan-archive.mrc-cbu.cam.ac.uk/"},
    {"pattern": r"\bmridata\.org\b|\bMRIData\b", "name": "mridata.org", "url": "https://mridata.org/"},
    {"pattern": r"\bNYU(?:\s+Langone)?\s+(?:brain|knee)\b", "name": "NYU fastMRI subset", "url": "https://fastmri.med.nyu.edu/"},
    {"pattern": r"\bLPBA40\b", "name": "LPBA40", "url": ""},
    {"pattern": r"\bHVSMR\b", "name": "HVSMR", "url": ""},
    {"pattern": r"\bLiTS\b|\bLITS\b", "name": "LiTS", "url": ""},
    {"pattern": r"\bImageNet\b", "name": "ImageNet", "url": "https://www.image-net.org/"},
    {"pattern": r"\bTCGA\b", "name": "TCGA", "url": "https://www.cancer.gov/tcga"},
    {"pattern": r"\bCAMELYON(?:16|17)?\b", "name": "CAMELYON", "url": "https://camelyon17.grand-challenge.org/"},
    {"pattern": r"\bPanNuke\b", "name": "PanNuke", "url": ""},
    {"pattern": r"\bMoNuSeg\b", "name": "MoNuSeg", "url": "https://monuseg.grand-challenge.org/"},
    {"pattern": r"\bCoNSeP\b", "name": "CoNSeP", "url": ""},
    {"pattern": r"\bPAIP\b(?=[^\n]{0,40}(?:challenge|dataset|WSI|histopathology))", "name": "PAIP", "url": ""},
    {"pattern": r"\bPANDA\b(?=[^\n]{0,40}(?:challenge|prostate|WSI))", "name": "PANDA", "url": ""},
    {"pattern": r"\bBCSS\b|\bBreast Cancer Semantic Segmentation\b", "name": "BCSS", "url": ""},
    {"pattern": r"\bBRACS\b", "name": "BRACS", "url": ""},
    {"pattern": r"\bDigestPath\b", "name": "DigestPath", "url": ""},
    {"pattern": r"\bNuCLS\b", "name": "NuCLS", "url": ""},
    {"pattern": r"\bLizard\b(?=[^\n]{0,30}(?:dataset|colorectal|nuclei))", "name": "Lizard", "url": ""},
]

# 发现新数据集时的句式模板（捕获组 1 为候选名称）
DATASET_DISCOVERY_PATTERNS: list[tuple[str, float]] = [
    (r"(?:the\s+)?([A-Z][A-Za-z0-9\-]+(?:[\s\-][A-Za-z0-9\-]+){0,4})\s+(?:dataset|data\s*set|database|benchmark|corpus)\b", 0.85),
    (r"\b(?:on|using|from|with|evaluated\s+on|trained\s+on|obtained\s+from|collected\s+from|provided\s+by)\s+(?:the\s+)?([A-Z][A-Za-z0-9\-]+(?:[\s\-][A-Za-z0-9\-]+){0,4})\s+(?:dataset|data\s*set|database)\b", 0.9),
    (r"\b([A-Z][A-Za-z0-9\-]+(?:[\s\-][A-Za-z0-9\-]+){0,3})\s+(?:challenge|competition)\b", 0.82),
    (r"\b(\d+\s+(?:subjects|patients|volunteers|scans|cases))\s+(?:from|in|of)\s+(?:the\s+)?([A-Z][A-Za-z0-9\-]+(?:[\s\-][A-Za-z0-9\-]+){0,3})\b", 0.78),
]

DATASET_SECTION_HEADINGS = re.compile(
    r"^(?:\d+(?:\.\d+)*\.?\s*)?"
    r"(DATA(?:\s+AND\s+CODE)?(?:SET)?S?|EXPERIMENTS?|EVALUATION|RESULTS|"
    r"MATERIALS?\s+AND\s+METHODS|IMPLEMENTATION\s+DETAILS?|STUDY\s+DESIGN|"
    r"SUBJECTS?\s+AND\s+METHODS|TRAINING\s+DETAILS?)\s*$",
    re.I | re.M,
)
NEXT_SECTION_HEADING = re.compile(
    r"^(?:\d+(?:\.\d+)*\.?\s*)?(INTRODUCTION|RELATED\s+WORK|METHODS?|CONCLUSION|DISCUSSION|REFERENCES|APPENDIX)\s*$",
    re.I | re.M,
)
DATASET_URL_RE = re.compile(
    r"https?://(?:[\w\-]+\.)?(?:zenodo\.org|synapse\.org|grand-challenge\.org|osf\.io|"
    r"openneuro\.org|mridata\.org|digital\.commons|figshare\.com|github\.com|"
    r"fastmri\.med\.nyu\.edu|adni\.loni\.usc\.edu|humanconnectome\.org|"
    r"brain-development\.org|ukbiobank\.ac\.uk|oasis-brains\.org)[^\s\]\)\"\'>,]+",
    re.I,
)
AVAILABLE_AT_RE = re.compile(
    r"(?:available\s+(?:at|from|on|via)|downloaded\s+from|accessed\s+(?:at|from)|obtained\s+from)\s+"
    r"(https?://[^\s\]\)\"\'>,]+)",
    re.I,
)

# 发现候选名称时的噪声词（小写）
DATASET_NOISE_TERMS = {
    "the", "this", "our", "a", "an", "training", "validation", "test", "testing",
    "public", "clinical", "retrospective", "prospective", "multi", "single",
    "whole", "full", "real", "simulated", "synthetic", "in", "vivo", "ex",
    "same", "following", "corresponding", "proposed", "new", "large", "small",
    "standard", "original", "additional", "separate", "independent", "internal",
    "external", "held", "out", "und", "ersampled", "sampled", "fully",
    "magnetic", "resonance", "imaging", "image", "mr", "mri", "k", "space",
    "brain", "knee", "cardiac", "cine", "multi", "contrast", "undersampled",
    "reconstruction", "network", "model", "method", "proposed", "baseline",
    "deep", "learning", "neural", "convolutional", "diffusion", "unrolled",
    "accelerated", "fast", "high", "quality", "low", "field", "whole", "body",
    "three", "dimensional", "d", "t1", "t2", "pd", "flair", "gre", "mse",
    "subject", "subjects", "patient", "patients", "volunteer", "volunteers",
    "scan", "scans", "case", "cases", "study", "data", "set", "dataset",
    "database", "benchmark", "corpus", "collection", "cohort", "sample",
    "samples", "split", "splits", "subset", "part", "portion", "region",
    "anatomical", "modality", "modalities", "sequence", "sequences",
    "open", "raw", "available", "obtained", "collected", "provided",
    "conducted", "performed", "experiments", "experiment", "ablation",
    "analysis", "evaluation", "comparison", "results", "images", "weighted",
    "initiative", "scratch", "dicom", "space", "weighted", "flair",
}

INVALID_DATASET_START = re.compile(
    r"^(?:for|an?|the|our|this|with|using|from|of|and|to|in|on|at|by|as|or|not|all|"
    r"each|both|such|same|other|several|various|multiple|single|results|images|"
    r"study|experiments|ablation|analysis|conducted|available|raw|obtained|"
    r"collected|provided|performed|evaluation|comparison|validation|training|"
    r"testing|test|brain|knee|cardiac|cine|dynamic|public|weighted|flair|"
    r"seven|low|field|mask|scratch|initiative|evaluations|beside|clinical|besides|"
    r"publicly|gen|nyu)\b",
    re.I,
)
INVALID_DATASET_CONTENT = re.compile(
    r"\b(provided\s+by|available|publicly|training|initiative|scratch|mask|"
    r"data\s+in|scan|images?\s+from|low[\-\s]field|brain\s+mri|knee\s+mri|"
    r"weighted|flair|evaluations?\s+on|experiments?\s+on|study\s+on|"
    r"clinical\s+brain|multi[\-\s]coil\s+knee)\b",
    re.I,
)

GITHUB_RE = re.compile(r"https?://github\.com/[\w\-\./]+", re.I)
DOI_RE = re.compile(r"(?:doi[:\s]*|https?://doi\.org/)(10\.\d{4,9}/[^\s\]\)\"\'>,]+)", re.I)
ARXIV_RE = re.compile(r"(?:arxiv[:\s]*|https?://arxiv\.org/abs/)(\d{4}\.\d{4,5}(?:v\d+)?)", re.I)
IEEE_DOI_RE = re.compile(r"10\.1109/[A-Z0-9\.]+", re.I)
YEAR_RE = re.compile(r"\b(20(?:1[5-9]|2[0-9]))\b")
EMAIL_RE = re.compile(r"\S+@\S+\.\S+")
IEEE_SUFFIX_RE = re.compile(
    r",?\s*(?:(?:Senior|Associate)\s+)?Member,?\s*IEEE|,\s*Fellow,?\s*IEEE|,\s*Graduate Student Member,?\s*IEEE",
    re.I,
)
AFFIL_LINE_RE = re.compile(
    r"^(?:[a-z]\d?(?:,\s*[a-z]\d?)*)?\s*(?:College|School|Department|University|Institute|Hospital|"
    r"Laboratory|Center|Centre|Lab\.|Ministry|USA|China|Germany|Faculty|Division|Medical Center)",
    re.I,
)
HEADER_RE = re.compile(
    r"(IEEE TRANSACTIONS|IEEE JOURNAL|VOL\.\s*\d|contents lists|sciencedirect|elsevier|"
    r"journal homepage|locate/media|authorized licensed|^\d{1,4}$|digital object|"
    r"Manuscript received|date of publication|Index Terms|Index terms)",
    re.I,
)
NON_NAME_WORDS = {
    "abstract", "introduction", "network", "model", "imaging", "reconstruction",
    "accelerated", "mri", "magnetic", "resonance", "using", "via", "for", "and",
    "the", "with", "from", "based", "learning", "diffusion", "deep", "fast",
    "member", "senior", "fellow", "ieee", "associate", "transformer", "prior",
    "feature", "scaling", "pearl", "camp-net", "modality", "sampling", "toward", "universal",
}


# ---------------------------------------------------------------------------
# 文本工具
# ---------------------------------------------------------------------------

def extract_text(pdf_path: Path, max_pages: int = 5) -> str:
    doc = fitz.open(pdf_path)
    parts = [doc[i].get_text("text") for i in range(min(max_pages, len(doc)))]
    doc.close()
    return "\n".join(parts)


def extract_dataset_text(pdf_path: Path, max_pages: int = 12) -> str:
    """数据集信息常出现在 Experiments 节，需读取更多页。"""
    doc = fitz.open(pdf_path)
    parts = [doc[i].get_text("text") for i in range(min(max_pages, len(doc)))]
    doc.close()
    return "\n".join(parts)


def clean_line(ln: str) -> str:
    return re.sub(r"\s+", " ", ln).strip()


def smart_title_case(s: str) -> str:
    if not s.isupper():
        return s.strip()
    small = {"for", "and", "of", "in", "on", "with", "via", "to", "a", "an", "the", "using", "from", "by"}
    out = []
    for i, w in enumerate(s.split()):
        lw = w.lower()
        if i > 0 and lw in small:
            out.append(lw)
        elif "-" in w:
            out.append("-".join(p.capitalize() for p in w.split("-")))
        else:
            out.append(w.capitalize())
    return " ".join(out)


def is_boilerplate(ln: str) -> bool:
    return not ln or bool(HEADER_RE.search(ln))


def is_elsevier_author_affil_marker(ln: str) -> bool:
    return bool(re.match(r"^[a-z]\d?\s*,", ln) or re.match(r"^,\s*[A-Z]", ln))


def is_affiliation_line(ln: str) -> bool:
    if is_elsevier_author_affil_marker(ln):
        return False
    if AFFIL_LINE_RE.search(ln):
        return True
    if re.match(r"^[a-z]\d?\s+[A-Z]", ln):
        return True
    if re.search(r"\b(university|institute|department|school|hospital|laboratory|center|centre)\b", ln, re.I):
        if not re.match(r"^[A-Z][a-z]+(?:\s+[A-Z][a-z]+)+$", ln):
            return True
    return False


def find_abstract_index(lines: list[str]) -> int:
    for i, ln in enumerate(lines[:80]):
        cl = clean_line(ln)
        if re.match(r"^abstract[\.\s\-—–:]", cl, re.I):
            return i
        if re.match(r"^ABSTRACT\b", cl):
            return i
        if re.sub(r"\s+", "", cl).upper() == "ABSTRACT":
            return i
        if re.match(r"^A\s+B\s+S\s+T\s+R\s+A\s+C\s+T", cl, re.I):
            return i
    return -1


# ---------------------------------------------------------------------------
# 标题提取
# ---------------------------------------------------------------------------

def title_from_filename(pdf_path: Path) -> str:
    name = pdf_path.stem
    if re.match(r"^\d+_paper$", name) or name.startswith("1-s2.0-"):
        return ""
    name = re.sub(r"^Liu\s*等\s*-\s*\d{4}\s*-\s*", "", name)
    title = re.sub(r"_+", " ", name).strip()
    for token in (
        "MRI", "MR", "MRF", "MRA", "CNN", "ViT", "INR", "KAN", "ECG", "HCP", "MMA",
        "SPIRiT", "PEARL", "PISCO", "PRIME", "HiMaC", "3D", "7T", "qMRI",
    ):
        title = re.sub(rf"\b{re.escape(token)}\b", token, title, flags=re.I)
    return title


def parse_elsevier_title(lines: list[str]) -> str:
    start = 0
    for i, ln in enumerate(lines[:30]):
        if "locate/media" in ln.lower():
            start = i + 1
            break
    parts = []
    for ln in lines[start : start + 8]:
        ln = clean_line(ln)
        if not ln or is_boilerplate(ln):
            continue
        if is_affiliation_line(ln):
            break
        if re.match(r"^[A-Z][a-zA-Z\-']+(?:\s+[A-Z][a-zA-Z\-']+){1,3}$", ln):
            break
        if re.search(r"\s[a-z]\s*,", ln):
            ln = re.sub(r"\s+[A-Z][a-zA-Z\-']+(?:\s+[A-Z][a-zA-Z\-']+)+$", "", ln).strip()
            if ln:
                parts.append(ln)
            break
        parts.append(ln)
    return clean_line(" ".join(parts))


def parse_ieee_title(lines: list[str]) -> str:
    parts, passed_header = [], False
    for ln in lines[:35]:
        ln = clean_line(ln)
        if not ln or is_boilerplate(ln):
            if re.match(r"^IEEE", ln, re.I):
                passed_header = True
            continue
        if re.match(r"^abstract\b", ln, re.I):
            break
        if re.search(r"\s[a-z]\s*,", ln) and re.search(r"[A-Z][a-z]+", ln):
            break
        if re.match(r"^[A-Z][a-zA-Z\-']+(?:,\s*[A-Z])", ln) and "," in ln:
            break
        parts.append(ln)
        if len(parts) >= 4:
            break
    return clean_line(" ".join(parts))


def parse_miccai_title(lines: list[str]) -> str:
    parts = []
    for ln in lines[:15]:
        ln = clean_line(ln)
        if not ln or is_boilerplate(ln):
            continue
        if re.search(r"\d{1,2},\d", ln) and re.search(r"[A-Z][a-z]+", ln):
            break
        if re.match(r"^[A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][a-z]+(?:\d|$)", ln):
            break
        parts.append(ln)
    return clean_line(" ".join(parts))


def parse_generic_title(lines: list[str]) -> str:
    parts = []
    for ln in lines[:20]:
        ln = clean_line(ln)
        if not ln or is_boilerplate(ln):
            continue
        if re.match(r"^abstract\b", ln, re.I) or is_affiliation_line(ln):
            break
        parts.append(ln)
        if len(" ".join(parts)) > 30:
            break
    title = clean_line(" ".join(parts))
    return smart_title_case(title) if title.isupper() else title


def guess_title(text: str, pdf_path: Path) -> str:
    fn_title = title_from_filename(pdf_path)
    if fn_title and len(fn_title) > 15:
        return fn_title
    lines = text.splitlines()
    name = pdf_path.name
    if name.startswith("1-s2.0-"):
        title = parse_elsevier_title(lines)
    elif re.match(r"^\d+_paper", name):
        title = parse_miccai_title(lines)
    elif pdf_path.parent.name in ("JBHI", "TMI"):
        title = parse_ieee_title(lines)
    else:
        title = parse_generic_title(lines)
    if title.isupper():
        title = smart_title_case(title)
    return title or fn_title


# ---------------------------------------------------------------------------
# 作者提取
# ---------------------------------------------------------------------------

def normalize_author_name(name: str) -> str:
    name = name.strip(" ,;.")
    name = re.sub(r"[\d\u00b9\u00b2\u00b3\u2070-\u2079\*†‡§∗]+", "", name)
    name = re.sub(r"\(\s*[A-Za-z]\s*\)", "", name)
    name = re.sub(r"\s+[a-z]$", "", name)
    name = re.sub(r"\s{2,}", " ", name).strip()
    name = IEEE_SUFFIX_RE.sub("", name).strip(" ,;.")
    name = re.sub(r"^and\s+", "", name, flags=re.I)
    return name


def is_valid_name(name: str) -> bool:
    if not name or len(name) < 3 or ":" in name or ";" in name:
        return False
    if EMAIL_RE.search(name):
        return False
    words = re.findall(r"[A-Za-z\-']+", name)
    if len(words) < 2 and "-" not in name:
        return False
    if len(words) > 6:
        return False
    low = name.lower()
    bad = sum(1 for w in words if w.lower() in NON_NAME_WORDS)
    if bad >= 2 or (bad >= 1 and len(words) <= 3):
        return False
    if re.search(
        r"\b(network|model|reconstruction|imaging|accelerated|diffusion|transformer|"
        r"using|prior|feature|scaling|pearl|camp-net|modality|sampling|toward|universal)\b",
        low,
    ):
        return False
    return bool(re.search(r"[A-Z][a-z]", name))


def split_author_block(block: str) -> list[str]:
    block = IEEE_SUFFIX_RE.sub("", block)
    block = re.sub(r"\s+", " ", block).strip()
    block = re.sub(r"\band\s+", ", ", block, flags=re.I)
    parts = [normalize_author_name(p) for p in block.split(",")]
    return [p for p in parts if is_valid_name(p)]


def strip_title_prefix_from_block(block: str, title: str) -> str:
    if not title or not block:
        return block
    t_words = [re.sub(r"[^\w\-]", "", w) for w in title.split()]
    b_words = block.split()
    for drop in range(min(len(t_words), len(b_words)), 0, -1):
        suffix = [w.lower() for w in t_words[-drop:]]
        prefix = [re.sub(r"[^\w\-]", "", w).lower() for w in b_words[:drop]]
        if suffix == prefix:
            return " ".join(b_words[drop:])
    return re.sub(r"^[A-Z][a-zA-Z\-]*-\s*", "", block)


def is_elsevier_author_line(ln: str) -> bool:
    if not ln or is_boilerplate(ln) or ln.strip() == "Medical Image Analysis":
        return False
    if is_affiliation_line(ln):
        return False
    if is_elsevier_author_affil_marker(ln):
        return True
    if re.search(r"\s[a-z]\d?\s*,", ln) and re.search(r"[A-Z][a-z]+", ln):
        return True
    if re.search(r"\s[a-z]\d?$", ln) and re.search(r"[A-Z][a-z]+\s+[A-Z]", ln):
        return True
    if re.match(r"^[A-Z][a-zA-Z\-'\u00c0-\u024f]+(?:\s+[A-Z][a-zA-Z\-'\u00c0-\u024f\.]+){1,4}$", ln):
        return True
    return False


def strip_elsevier_affiliation_markers(block: str) -> str:
    block = re.sub(r"\s+[a-z]\d?(?=,)", "", block)
    block = re.sub(r",\s*[a-z]\d?(?=,|\s|$)", ",", block)
    block = re.sub(r"\s*,?\s*[\*†‡§∗]+", "", block)
    return re.sub(r"\s{2,}", " ", block).strip(" ,;")


def elsevier_content_start(lines: list[str]) -> int:
    for i, ln in enumerate(lines[:35]):
        if "locate/media" in ln.lower():
            return i + 1
    for i, ln in enumerate(lines[:35]):
        cl = clean_line(ln)
        if not cl or is_boilerplate(cl):
            continue
        if re.match(r"^Medical Image Analysis\s+\d", cl, re.I):
            continue
        if re.search(r"available online|all rights|1361-8415", cl, re.I):
            continue
        return i
    return 0


def extract_ieee_authors(text: str, title: str) -> str:
    lines = [clean_line(ln) for ln in text.splitlines()]
    abs_idx = find_abstract_index(lines)
    if abs_idx < 0:
        abs_idx = 45

    author_lines = []
    for i, ln in enumerate(lines[:abs_idx]):
        if not ln or is_boilerplate(ln) or is_affiliation_line(ln):
            if is_affiliation_line(ln):
                break
            continue
        if re.match(r"^Manuscript received", ln, re.I):
            break
        has_signal = (
            "," in ln
            or re.search(r"\band\s+[A-Z][a-z]", ln)
            or re.search(r"Member,?\s*IEEE|Fellow,?\s*IEEE", ln, re.I)
        )
        is_lone_name = bool(
            re.match(r"^[A-Z][a-zA-Z\-'\.\u00c0-\u024f]+(?:\s+[A-Z][a-zA-Z\-'\.\u00c0-\u024f]+)+$", ln)
            and i + 1 < abs_idx
            and (lines[i + 1].startswith(",") or "IEEE" in lines[i + 1])
        )
        if has_signal or is_lone_name:
            author_lines.append(ln)

    block = strip_title_prefix_from_block(" ".join(author_lines), title)
    return "; ".join(split_author_block(block))


def extract_elsevier_authors(text: str, title: str) -> str:
    lines = [clean_line(ln) for ln in text.splitlines()]
    abs_idx = find_abstract_index(lines)
    if abs_idx < 0:
        abs_idx = 40

    start = elsevier_content_start(lines)
    phase = "skip_title"
    author_lines: list[str] = []

    for ln in lines[start:abs_idx]:
        if not ln or is_boilerplate(ln):
            continue
        if is_affiliation_line(ln):
            break
        if phase == "skip_title":
            inline = re.search(
                r"\s((?:[A-Z][a-zA-Z\-']+(?:\s+[A-Z][a-zA-Z\-']+)+\s+[a-z]\d?(?:,|\s).+))$",
                ln,
            )
            if inline:
                phase = "authors"
                author_lines.append(inline.group(1))
                continue
            if is_elsevier_author_line(ln):
                phase = "authors"
                author_lines.append(ln)
            continue
        if is_elsevier_author_line(ln):
            author_lines.append(ln)
        else:
            break

    block = strip_elsevier_affiliation_markers(" ".join(author_lines))
    return "; ".join(split_author_block(block))


def extract_miccai_authors(text: str, title: str) -> str:
    lines = [clean_line(ln) for ln in text.splitlines()]
    abs_idx = find_abstract_index(lines)
    if abs_idx < 0:
        abs_idx = 20

    for ln in lines[:abs_idx]:
        if not ln or is_boilerplate(ln):
            continue
        tw = [w.lower() for w in re.findall(r"[A-Za-z0-9\-]+", title) if len(w) > 2]
        if tw and sum(1 for w in tw if w in ln.lower()) >= max(2, len(tw) * 0.5):
            continue
        if re.search(r"\d{1,2}(?:,\d{1,2})+", ln) and re.search(r"[A-Z][a-z]+", ln):
            names = split_author_block(re.sub(r"\d{1,2}(?:,\d{1,2})*", "", ln))
            if names:
                return "; ".join(names)
        if re.search(r",\s*and\s+[A-Z]", ln) and re.search(r"[A-Z][a-z]+", ln):
            names = split_author_block(ln)
            if names:
                return "; ".join(names)
    return ""


def extract_isbi_authors(text: str, title: str) -> str:
    lines = [clean_line(ln) for ln in text.splitlines()]
    abs_idx = find_abstract_index(lines)
    if abs_idx < 0:
        abs_idx = 25

    names: list[str] = []
    for ln in lines[:abs_idx]:
        if not ln or is_boilerplate(ln):
            continue
        tw = [w.lower() for w in re.findall(r"[A-Za-z0-9\-]+", title) if len(w) > 2]
        if tw and sum(1 for w in tw if w in ln.lower()) >= max(2, len(tw) * 0.55):
            continue
        if is_affiliation_line(ln):
            break
        if re.match(r"^[A-Z][a-zA-Z\-'\u00c0-\u024f]+(?:\s+[A-Z][a-zA-Z\-'\u00c0-\u024f\.]+)+\d+$", ln):
            n = normalize_author_name(ln)
            if is_valid_name(n):
                names.append(n)
        elif re.match(r"^[A-Z][a-zA-Z\-'\u00c0-\u024f]+(?:\s+[A-Z][a-zA-Z\-'\u00c0-\u024f]+)+$", ln):
            n = normalize_author_name(ln)
            if is_valid_name(n):
                names.append(n)
        elif re.search(r",\s*and\s+", ln, re.I):
            names.extend(split_author_block(ln))
            break
    return "; ".join(names)


def extract_authors(text: str, pdf_path: Path, title: str) -> str:
    folder = pdf_path.parent.name
    name = pdf_path.name

    if name.startswith("1-s2.0-"):
        result = extract_elsevier_authors(text, title)
    elif re.match(r"^\d+_paper", name):
        result = extract_miccai_authors(text, title)
    elif folder in ("JBHI", "TMI"):
        result = extract_ieee_authors(text, title)
    elif folder == "ISBI 2026":
        result = extract_isbi_authors(text, title)
    else:
        result = extract_ieee_authors(text, title) or extract_elsevier_authors(text, title)

    return result or extract_miccai_authors(text, title)


# ---------------------------------------------------------------------------
# 其他字段
# ---------------------------------------------------------------------------

def xl_safe(value: Any) -> Any:
    """去除 openpyxl 不允许的控制字符。"""
    if value is None or not isinstance(value, str):
        return value
    return _ILLEGAL_XL_RE.sub("", value)


def guess_venue(text: str, pdf_path: Path) -> str:
    """从 PDF 首页识别期刊/会议名称，失败时回退到文件夹名。"""
    url = guess_urls(text)
    doi_venue = {
        r"10\.1109/TMI": "IEEE Transactions on Medical Imaging (TMI)",
        r"10\.1109/JBHI": "IEEE Journal of Biomedical and Health Informatics (JBHI)",
        r"10\.1109/TCSVT": "IEEE Transactions on Circuits and Systems for Video Technology (TCSVT)",
        r"10\.1109/TIP": "IEEE Transactions on Image Processing (TIP)",
        r"10\.1109/isbi": "IEEE ISBI",
        r"10\.1109/ISPCT": "IEEE ISPCT",
        r"10\.1016/j\.media": "Medical Image Analysis",
        r"10\.1007/": "Springer (MICCAI/期刊)",
        r"10\.3389/": "Frontiers",
        r"arxiv": "arXiv",
    }
    for pattern, name in doi_venue.items():
        if url and re.search(pattern, url, re.I):
            return name

    head = text[:4000]
    for pattern, name in VENUE_PATTERNS:
        if re.search(pattern, head, re.I):
            return name
    m = re.search(
        r"(IEEE\s+(?:Transactions on|Journal of)\s+[A-Za-z\s&]+?)(?:,|\s+VOL\.)",
        head,
        re.I,
    )
    if m:
        return re.sub(r"\s+", " ", m.group(1)).strip()
    m = re.search(r"Published in\s+(.{5,80}?)(?:\.|,|\n)", head, re.I)
    if m:
        return clean_line(m.group(1))
    folder = pdf_path.parent.name
    return FOLDER_VENUE.get(folder, folder)


def _doi_from_url(url: str) -> str:
    m = re.search(r"doi\.org/(.+)$", url, re.I)
    return m.group(1).rstrip(".") if m else ""


def fetch_citation_count(url: str, *, timeout: float = 8.0) -> str:
    """通过 Semantic Scholar API 查询引用次数（需网络）。"""
    doi = _doi_from_url(url)
    if not doi:
        return ""
    api = f"https://api.semanticscholar.org/graph/v1/paper/DOI:{doi}?fields=citationCount"
    req = urllib.request.Request(api, headers={"User-Agent": "paper-metadata-extractor/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = json.loads(resp.read().decode())
        count = data.get("citationCount")
        return str(count) if count is not None else ""
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, json.JSONDecodeError, KeyError):
        return ""


def guess_year(text: str, pdf_path: Path) -> str:
    m = re.search(r"VOL\.\s*\d+,\s*NO\.\s*\d+,\s*[A-Z]+\s+(\d{4})", text, re.I)
    if m:
        return m.group(1)
    m = re.search(r"Medical Image Analysis\s+\d+\s*\((\d{4})\)", text, re.I)
    if m:
        return m.group(1)
    m = re.search(r"-\s*(\d{4})\s*-", pdf_path.name)
    if m:
        return m.group(1)
    m = re.search(r"S13618415(\d{4})", pdf_path.name)
    if m:
        y = int(m.group(1))
        if 2015 <= y <= 2030:
            return str(y)
    years = YEAR_RE.findall(text[:5000])
    return Counter(years).most_common(1)[0][0] if years else ""


def guess_urls(text: str) -> str:
    m = DOI_RE.search(text)
    if m:
        return f"https://doi.org/{m.group(1).rstrip('.')}"
    m = IEEE_DOI_RE.search(text)
    if m:
        return f"https://doi.org/{m.group(0).rstrip('.')}"
    m = ARXIV_RE.search(text)
    if m:
        return f"https://arxiv.org/abs/{m.group(1)}"
    return ""


def guess_github(text: str) -> tuple[str, str]:
    urls = GITHUB_RE.findall(text)
    if urls:
        return "是", urls[0].rstrip(").,;]")
    if re.search(r"\bcode\s+(?:is|will be)\s+(?:available|released)\b", text, re.I):
        return "待公开", ""
    if re.search(r"\b(?:source code|github|code repository|open.?source)\b", text, re.I):
        return "文中提及", ""
    return "否", ""


# ---------------------------------------------------------------------------
# 数据集提取（分层）
# ---------------------------------------------------------------------------

@dataclass
class DatasetMention:
    name: str
    url: str = ""
    source: str = "unknown"  # known | section | pattern | url | table
    confidence: float = 0.5

    def key(self) -> str:
        n = re.sub(r"\s+", " ", self.name.strip().lower())
        n = re.sub(r"\s+(knee|brain|cardiac|cine|flair|weighted|subset)$", "", n)
        return n


def _prepare_dataset_text(text: str) -> str:
    """去除 PDF 换行连字符，便于模式匹配。"""
    text = re.sub(r"-\s*\n\s*", "", text)
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_dataset_name(name: str) -> str:
    name = re.sub(r"\s+", " ", name.strip(" .,;:-"))
    name = re.sub(r"\s+(dataset|data\s*set|database|benchmark|corpus|challenge)$", "", name, flags=re.I)
    return name.strip()


def _is_valid_discovered_name(name: str) -> bool:
    name = _normalize_dataset_name(name)
    if not name or len(name) < 3 or len(name) > 60:
        return False
    if INVALID_DATASET_START.match(name):
        return False
    if INVALID_DATASET_CONTENT.search(name):
        return False
    if name.lower() in DATASET_NOISE_TERMS:
        return False
    words = re.findall(r"[A-Za-z0-9\-]+", name)
    if not words or len(words) > 5:
        return False
    if all(w.lower() in DATASET_NOISE_TERMS for w in words):
        return False
    if len(words) == 1 and words[0].lower() in {"mri", "mr", "fmri", "dmri", "qmri", "hcp"}:
        return False
    if not re.match(r"^[A-Z0-9]", name):
        return False
    for word in name.split():
        wl = word.lower()
        if wl in {"of", "and", "the", "for", "on", "in", "at", "by", "we", "use", "evaluate", "evaluated"}:
            return False
        if word[0].islower() and wl not in {"knee", "brain", "cardiac", "cine"}:
            return False
    if re.search(r"\b(network|model|module|block|loss|algorithm|framework|architecture|images|results|experiments|study|analysis|ablation)\b", name, re.I):
        return False
    # 非缩写名至少 2 个实词，且不能全是解剖/序列类型词
    has_acronym = bool(re.search(r"[A-Z]{2,}[A-Za-z0-9]*|\b[A-Z][a-z]+[A-Z]", name))
    content_words = [w for w in words if w.lower() not in DATASET_NOISE_TERMS]
    if not has_acronym and len(content_words) < 2:
        return False
    return True


def _extract_dataset_sections(text: str) -> str:
    """拼接 Dataset / Experiments / Methods 等相关章节文本。"""
    chunks: list[str] = []
    for m in DATASET_SECTION_HEADINGS.finditer(text):
        start = m.end()
        tail = text[start : start + 6000]
        end_m = NEXT_SECTION_HEADING.search(tail)
        chunk = tail[: end_m.start()] if end_m else tail[:4000]
        chunks.append(chunk)
    return "\n".join(chunks)


def _match_known_datasets(text: str) -> list[DatasetMention]:
    found: list[DatasetMention] = []
    seen: set[str] = set()
    for entry in KNOWN_DATASETS:
        if re.search(entry["pattern"], text, re.I):
            key = entry["name"].lower()
            if key not in seen:
                seen.add(key)
                found.append(DatasetMention(
                    name=entry["name"],
                    url=entry.get("url", ""),
                    source="known",
                    confidence=0.95,
                ))
    return found


def _discover_datasets_by_patterns(text: str, source: str, boost: float = 0.0) -> list[DatasetMention]:
    found: list[DatasetMention] = []
    seen: set[str] = set()
    for pattern, base_conf in DATASET_DISCOVERY_PATTERNS:
        for m in re.finditer(pattern, text, re.I):
            groups = [g for g in m.groups() if g]
            candidate = groups[-1] if groups else ""
            candidate = _normalize_dataset_name(candidate)
            if not _is_valid_discovered_name(candidate):
                continue
            key = candidate.lower()
            if key in seen:
                continue
            seen.add(key)
            found.append(DatasetMention(
                name=candidate,
                source=source,
                confidence=min(1.0, base_conf + boost),
            ))
    return found


def _extract_table_dataset_rows(text: str) -> list[DatasetMention]:
    """Table I / Dataset 列：| fastMRI | 或 Dataset  xxx  #subjects"""
    found: list[DatasetMention] = []
    seen: set[str] = set()
    patterns = [
        r"(?:^|\n)\s*(?:Dataset|Data\s*set)\s*[:|\|]\s*([A-Za-z0-9][A-Za-z0-9\-\s]{2,40}?)(?:\s*[\|,\t]|\s+\d|\s*$)",
        r"(?:^|\n)\s*TABLE\s+[IVX\d]+[^\n]*\n(?:[^\n]*\n){0,3}[^\n]*\b([A-Z][A-Za-z0-9\-]+(?:\s+[A-Za-z0-9\-]+){0,3})\b[^\n]*\d+\s*(?:subjects|patients|scans|cases)",
        r"(?:^|\n)\s*[•\-\*]\s*([A-Z][A-Za-z0-9\-]+(?:\s+[A-Za-z0-9\-]+){0,4})\s*(?:dataset|data\s*set|database)\b",
    ]
    for pat in patterns:
        for m in re.finditer(pat, text, re.I | re.M):
            candidate = _normalize_dataset_name(m.group(1))
            if not _is_valid_discovered_name(candidate):
                continue
            key = candidate.lower()
            if key in seen:
                continue
            seen.add(key)
            found.append(DatasetMention(name=candidate, source="table", confidence=0.7))
    return found


def _extract_urls_from_text(text: str) -> list[str]:
    urls = DATASET_URL_RE.findall(text) + [m.group(1) for m in AVAILABLE_AT_RE.finditer(text)]
    cleaned = []
    for u in urls:
        u = u.rstrip(").,;]")
        if u not in cleaned:
            cleaned.append(u)
    return cleaned


def _associate_urls(mentions: list[DatasetMention], text: str) -> None:
    """在数据集名称附近窗口内搜索 URL，并填充已知目录默认链接。"""
    all_urls = _extract_urls_from_text(text)
    known_url_map = {e["name"].lower(): e.get("url", "") for e in KNOWN_DATASETS}

    for m in mentions:
        if not m.url and m.key() in known_url_map:
            m.url = known_url_map[m.key()]
        if m.url:
            continue
        # 在全文搜索名称 ±120 字符窗口内的 URL
        for match in re.finditer(re.escape(m.name.split("(")[0].strip()), text, re.I):
            window = text[max(0, match.start() - 120) : match.end() + 200]
            local_urls = _extract_urls_from_text(window)
            if local_urls:
                m.url = local_urls[0]
                break
        # 仍未找到则尝试域名关键词匹配
        if not m.url:
            name_token = m.name.split()[0].lower()
            for u in all_urls:
                if name_token in u.lower():
                    m.url = u
                    break


def _merge_dataset_mentions(mentions: list[DatasetMention]) -> list[DatasetMention]:
    """按 canonical key 合并，保留最高置信度及已知来源优先。"""
    merged: dict[str, DatasetMention] = {}
    for m in mentions:
        key = m.key()
        if key not in merged:
            merged[key] = m
            continue
        prev = merged[key]
        if m.confidence > prev.confidence or (m.source == "known" and prev.source != "known"):
            merged[key] = DatasetMention(
                name=prev.name if prev.source == "known" else m.name,
                url=m.url or prev.url,
                source=m.source if m.source == "known" else prev.source,
                confidence=max(m.confidence, prev.confidence),
            )
        elif not prev.url and m.url:
            prev.url = m.url
    return sorted(merged.values(), key=lambda x: (-x.confidence, x.name.lower()))


def extract_datasets(text: str) -> tuple[str, str]:
    """
    分层提取数据集。

    1. 已知目录匹配（高置信）
    2. Dataset/Experiments 章节内模式发现（+0.05 置信加成）
    3. 全文模式发现
    4. 表格/列表行解析
    5. URL 关联与去重

    Returns
    -------
    (datasets_str, dataset_urls_str) — 分号分隔，兼容 Excel 列格式
    """
    text = _prepare_dataset_text(text)
    section_text = _extract_dataset_sections(text)
    mentions: list[DatasetMention] = []

    mentions.extend(_match_known_datasets(text))
    if section_text:
        mentions.extend(_discover_datasets_by_patterns(section_text, "section", boost=0.05))
        mentions.extend(_extract_table_dataset_rows(section_text))
    else:
        # 无明确章节时，仅在摘要后全文中做保守发现
        body = text[text.lower().find("abstract") + 8 :] if "abstract" in text.lower() else text
        mentions.extend(_discover_datasets_by_patterns(body[:12000], "pattern"))
    mentions.extend(_extract_table_dataset_rows(text))

    merged = _merge_dataset_mentions(mentions)
    _associate_urls(merged, text)

    # 过滤：发现项需更高置信；已知项始终保留
    final = [m for m in merged if m.source == "known" or m.confidence >= 0.72]

    names = [m.name for m in final]
    url_parts = [f"{m.name}: {m.url}" for m in final if m.url]
    return "; ".join(names), "; ".join(url_parts)


def guess_datasets(text: str) -> tuple[str, str]:
    """向后兼容别名。"""
    return extract_datasets(text)


def guess_summary(text: str) -> str:
    """提取摘要片段（兼容 IEEE 同行 Abstract—、Elsevier 分行 A B S T R A C T 等格式）。"""
    patterns = [
        r"abstract[\s\-–—:]+(.+?)(?=\n\s*(?:index terms|keywords|\d+\.\s*introduction|I\.\s+INTRODUCTION))",
        r"abstract[\s\-–—:]*\n(.+?)(?=\n\s*(?:index terms|keywords|\d+\.\s*introduction|I\.\s+INTRODUCTION))",
        r"A\s+B\s+S\s+T\s+R\s+A\s+C\s+T\s*\n(.+?)(?=\n\s*(?:keywords|key words|index terms|\d+\.\s*introduction|1\.\s*introduction|highlights|graphical abstract))",
        r"ABSTRACT[\s\-–—:]*\n(.+?)(?=\n\s*(?:KEYWORDS|INDEX TERMS|INTRODUCTION|\d+\.\s*Introduction))",
        r"abstract[\s\-–—:]+(.{80,800})",
        r"abstract[\s\-–—:]*\n(.{80,800})",
        r"A\s+B\s+S\s+T\s+R\s+A\s+C\s+T\s*\n(.{80,800})",
    ]
    for pat in patterns:
        m = re.search(pat, text[:8000], re.I | re.S)
        if m:
            s = re.sub(r"\s+", " ", m.group(1)).strip()
            s = re.sub(r"-\s+", "", s)
            s = re.sub(r"^[\—\-–:\s]+", "", s)
            if len(s) >= 60:
                return (s[:197] + "...") if len(s) > 200 else s
    return ""


# ---------------------------------------------------------------------------
# 公开 API
# ---------------------------------------------------------------------------

def extract_paper_info(pdf_path: Path, *, fetch_citations: bool = False) -> dict[str, Any]:
    """从单篇 PDF 提取全部元数据字段。"""
    folder = pdf_path.parent.name
    text = extract_text(pdf_path)
    dataset_text = extract_dataset_text(pdf_path)
    title = guess_title(text, pdf_path)
    has_code, code_url = guess_github(text)
    datasets, dataset_urls = extract_datasets(dataset_text + "\n" + text[:8000])
    url = guess_urls(text)
    citations = fetch_citation_count(url) if fetch_citations and url else ""
    return {
        "title": title,
        "summary": guess_summary(text),
        "authors": extract_authors(text, pdf_path, title),
        "venue": guess_venue(text, pdf_path),
        "year": guess_year(text, pdf_path),
        "citations": citations,
        "url": url,
        "has_code": has_code,
        "code_url": code_url,
        "datasets": datasets,
        "dataset_urls": dataset_urls,
        "notes": f"来源: {folder}/{pdf_path.name}",
    }


@dataclass
class FillReport:
    """Excel 填充统计报告。"""
    total: int = 0
    field_counts: dict[str, int] = field(default_factory=dict)
    rows: list[dict[str, Any]] = field(default_factory=list)

    def record(self, info: dict[str, Any]) -> None:
        self.total += 1
        mapping = {
            "title": info.get("title"),
            "summary": info.get("summary"),
            "authors": info.get("authors"),
            "venue": info.get("venue"),
            "year": info.get("year"),
            "citations": info.get("citations"),
            "url": info.get("url"),
            "has_code": info.get("has_code"),
            "code_url": info.get("code_url"),
            "datasets": info.get("datasets"),
            "dataset_urls": info.get("dataset_urls"),
            "notes": info.get("notes"),
        }
        self.rows.append(mapping)
        for key, val in mapping.items():
            if val not in (None, "", "否"):
                self.field_counts[key] = self.field_counts.get(key, 0) + 1

    def format_summary(self) -> str:
        labels = {
            "title": "论文题目",
            "summary": "摘要(C列)",
            "authors": "作者",
            "venue": "期刊/会议",
            "year": "发表时间",
            "citations": "谷歌引用率",
            "url": "论文网址",
            "has_code": "是否有代码",
            "code_url": "代码地址",
            "datasets": "所用数据集",
            "dataset_urls": "数据集地址",
            "notes": "备注",
        }
        lines = [f"共处理 {self.total} 篇论文", "", "字段填充率："]
        for key, label in labels.items():
            n = self.field_counts.get(key, 0)
            if key == "has_code":
                n = sum(1 for r in self.rows if r.get("has_code") not in (None, ""))
            pct = (n / self.total * 100) if self.total else 0
            lines.append(f"  {label}: {n}/{self.total} ({pct:.0f}%)")
        return "\n".join(lines)


def _write_info_row(ws, row: int, idx: int, info: dict[str, Any]) -> None:
    ws.cell(row=row, column=1, value=idx)
    ws.cell(row=row, column=2, value=xl_safe(info["title"]))
    ws.cell(row=row, column=3, value=xl_safe(info["summary"]))
    ws.cell(row=row, column=4, value=xl_safe(info["authors"]))
    ws.cell(row=row, column=5, value=xl_safe(info["venue"]))
    ws.cell(row=row, column=6, value=xl_safe(info["year"]))
    ws.cell(row=row, column=7, value=xl_safe(info.get("citations", "")))
    ws.cell(row=row, column=8, value=xl_safe(info["url"]))
    ws.cell(row=row, column=9, value=xl_safe(info["has_code"]))
    ws.cell(row=row, column=10, value=xl_safe(info["code_url"]))
    ws.cell(row=row, column=11, value=xl_safe(info["datasets"]))
    ws.cell(row=row, column=12, value=xl_safe(info["dataset_urls"]))
    ws.cell(row=row, column=13, value=xl_safe(info["notes"]))
    ws.cell(row=row, column=14, value="")


def _clear_rows(ws, start_row: int, end_row: int, num_cols: int = 14) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            # 模板 A 列可能含 =A{n-1}+1 公式，需用空字符串覆盖
            cell.value = "" if col == 1 and getattr(cell, "data_type", None) == "f" else None
            if getattr(cell, "data_type", None) == "f" and col != 1:
                cell.value = None


def fill_summary_xlsx(
    xlsx_path: Path,
    papers_root: Path,
    *,
    sheet_title: str = "MR加速成像相关文献整理",
    start_row: int = 5,
    folder_order: list[str] | None = None,
    copy_from: Path | None = None,
    fetch_citations: bool = False,
    citation_delay: float = 0.5,
    update_authors_only: bool = False,
    update_datasets_only: bool = False,
) -> tuple[int, FillReport]:
    """
    扫描 papers_root 下所有 PDF，写入 summary.xlsx。

    Parameters
    ----------
    xlsx_path : Excel 文件路径
    papers_root : 论文 PDF 根目录（含 TMI/Media/JBHI 等子文件夹）
    copy_from : 若指定，先将模板复制到 xlsx_path
    fetch_citations : 是否通过 Semantic Scholar 查询引用次数（G 列）
    update_authors_only : 若为 True，仅更新 D 列（作者），保留已有其他列
    update_datasets_only : 若为 True，仅更新 K/L 列（数据集），保留已有其他列

    Returns
    -------
    (写入论文数量, FillReport)
    """
    if copy_from is not None:
        shutil.copy2(copy_from, xlsx_path)

    order = folder_order or DEFAULT_FOLDER_ORDER
    pdfs = sorted(
        papers_root.rglob("*.pdf"),
        key=lambda p: (
            order.index(p.parent.name) if p.parent.name in order else 99,
            p.name.lower(),
        ),
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    if start_row > 2:
        ws["A1"] = sheet_title

    report = FillReport()
    for idx, pdf in enumerate(pdfs, start=1):
        row = start_row + idx - 1
        info = extract_paper_info(pdf, fetch_citations=fetch_citations)
        if fetch_citations and info["url"] and citation_delay > 0:
            time.sleep(citation_delay)
        report.record(info)
        if update_authors_only:
            ws.cell(row=row, column=4, value=xl_safe(info["authors"]) or None)
        elif update_datasets_only:
            ws.cell(row=row, column=11, value=xl_safe(info["datasets"]) or None)
            ws.cell(row=row, column=12, value=xl_safe(info["dataset_urls"]) or None)
        else:
            _write_info_row(ws, row, idx, info)

    if not update_authors_only and not update_datasets_only:
        last_data_row = start_row + len(pdfs) - 1
        if ws.max_row > last_data_row:
            ws.delete_rows(last_data_row + 1, ws.max_row - last_data_row)

    wb.save(xlsx_path)
    return len(pdfs), report


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="从 PDF 提取论文元数据并写入 Excel")
    parser.add_argument("--root", type=Path, required=True, help="论文 PDF 根目录")
    parser.add_argument("--xlsx", type=Path, required=True, help="summary.xlsx 路径")
    parser.add_argument("--copy-from", type=Path, default=None, help="从模板复制后再写入")
    parser.add_argument("--title", type=str, default="MR加速成像相关文献整理", help="Sheet 标题(A1，start-row>2 时写入)")
    parser.add_argument("--start-row", type=int, default=5, help="数据起始行（默认 5；表头在第 1 行时用 2）")
    parser.add_argument("--fetch-citations", action="store_true", help="查询 Semantic Scholar 引用次数")
    parser.add_argument("--authors-only", action="store_true", help="仅更新作者列")
    parser.add_argument("--datasets-only", action="store_true", help="仅更新数据集列")
    args = parser.parse_args()
    n, report = fill_summary_xlsx(
        args.xlsx,
        args.root,
        sheet_title=args.title,
        start_row=args.start_row,
        copy_from=args.copy_from,
        fetch_citations=args.fetch_citations,
        update_authors_only=args.authors_only,
        update_datasets_only=args.datasets_only,
    )
    print(f"已处理 {n} 篇论文 → {args.xlsx}")
    print(report.format_summary())
